#!/usr/bin/env python3
"""
Читает Товары.xlsx в корне репозитория и пишет файлы для импорта в NocoBase:

- import/products-nocobase.csv / .xlsx — лист **активный при сохранении** (как раньше: обычно «Data» с товарами).
- import/nocobase-subscription_branding.csv / .xlsx — первая колонка **ID** (пусто = новая запись), далее ключи полей, лист **Sheet1**.
- import/nocobase-subscription_branding-RU-titles.xlsx — то же, подписи полей по-русски (кроме **ID**).

Зависимость: pip install openpyxl

Колонки экспорта NocoBase могут называться «Целое число» вместо grantDays — скрипт это учитывает.
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Товары.xlsx"
OUT_CSV = ROOT / "import" / "products-nocobase.csv"
OUT_XLSX = ROOT / "import" / "products-nocobase.xlsx"
# Имя с префиксом коллекции — чтобы в NocoBase не перепутали с импортом «Товары» (там другие заголовки).
OUT_BRANDING_CSV = ROOT / "import" / "nocobase-subscription_branding.csv"
OUT_BRANDING_XLSX = ROOT / "import" / "nocobase-subscription_branding.xlsx"
OUT_BRANDING_XLSX_RU = ROOT / "import" / "nocobase-subscription_branding-RU-titles.xlsx"

BRANDING_SHEET = "subscription_branding"

# NocoBase часто требует колонку ID первой (пустая при создании записи).
BRANDING_FIELD_KEYS = [
    "ID",
    "subscriptionTitle",
    "supportUrl",
    "profileUrl",
    "announcement",
    "active",
]

BRANDING_RU_TITLES = [
    "ID",
    "Заголовок подписки",
    "URL поддержки",
    "URL профиля",
    "Объявление",
    "Активен",
]

# Импорт в NocoBase ожидает те же заголовки, что в «Экспорт Excel» из UI (не имена полей API).
NOCOBASE_IMPORT_HEADERS = [
    "ID",
    "code",
    "Целое число",
    "productType",
    "title",
    "sortOrder",
    "active",
    "serverId",
]


def internal_to_nocobase_import_rows(rows: list[dict[str, object]]) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Возвращает (строки для Excel, строки для CSV). В Excel — типы как в экспорте; в CSV active — строки True/False."""
    xlsx_rows: list[dict[str, object]] = []
    csv_rows: list[dict[str, object]] = []
    for r in rows:
        active_raw = r.get("active", "false")
        is_act = str(active_raw).strip().lower() in ("1", "true", "yes", "да")
        try:
            gd = int(float(r.get("grantDays", 0)))
        except (TypeError, ValueError):
            gd = 0
        try:
            so = int(float(r.get("sortOrder", 0)))
        except (TypeError, ValueError):
            so = 0
        sid = str(r.get("serverId") or "").strip()
        base = {
            "ID": None,
            "code": r.get("code", ""),
            "Целое число": gd,
            "productType": r.get("productType", ""),
            "title": r.get("title", ""),
            "sortOrder": so,
            "serverId": sid if sid else None,
        }
        xlsx_rows.append({**base, "active": is_act})
        csv_rows.append({**base, "active": "True" if is_act else "False"})
    return xlsx_rows, csv_rows


def _find_sheet(wb: object, title: str) -> object | None:
    want = title.strip().lower()
    for ws in wb.worksheets:
        if str(ws.title).strip().lower() == want:
            return ws
    return None


def export_subscription_branding(wb: object) -> int:
    """Лист subscription_branding → CSV/XLSX для NocoBase. Возвращает число строк данных."""
    ws = _find_sheet(wb, BRANDING_SHEET)
    if ws is None:
        print(f"(нет листа «{BRANDING_SHEET}» в Товары.xlsx — пропуск брендинга)")
        return 0

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"Лист «{BRANDING_SHEET}» пуст")
        return 0

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    lowered = [h.lower() for h in header]

    def idx(*cands: str) -> int | None:
        for cand in cands:
            c = cand.lower()
            if c in lowered:
                return lowered.index(c)
        return None

    def idx_fuzzy(*needles: str) -> int | None:
        for j, h in enumerate(header):
            hl = h.lower()
            for n in needles:
                if n.lower() in hl:
                    return j
        return None

    i_title = idx("subscriptiontitle") or idx_fuzzy("заголовок подписки", "subscriptiontitle")
    if i_title is None:
        i_title = idx("title")
    i_sup = idx("supporturl") or idx_fuzzy("url поддержки", "поддержк")
    i_prof = idx("profileurl") or idx_fuzzy("url профиля", "профил")
    i_ann = idx("announcement") or idx_fuzzy("объявлен")
    i_act = idx("active")

    if i_title is None:
        print(
            f"Лист «{BRANDING_SHEET}»: нет колонки subscriptionTitle (или title / «Заголовок»).",
            file=sys.stderr,
        )
        print("Заголовки:", header, file=sys.stderr)
        return 0

    data_rows: list[dict[str, object]] = []
    for r in rows[1:]:
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        def cell(i: int | None) -> str:
            if i is None or i >= len(r):
                return ""
            v = r[i]
            if v is None:
                return ""
            return str(v).strip()

        title = cell(i_title)
        if not title:
            continue
        active = True
        if i_act is not None:
            v = r[i_act] if i_act < len(r) else None
            if isinstance(v, bool):
                active = v
            elif v is not None:
                active = str(v).strip().lower() in ("1", "true", "yes", "да")

        data_rows.append(
            {
                "ID": None,
                "subscriptionTitle": title,
                "supportUrl": cell(i_sup),
                "profileUrl": cell(i_prof),
                "announcement": cell(i_ann),
                "active": "True" if active else "False",
            }
        )

    OUT_BRANDING_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_BRANDING_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=BRANDING_FIELD_KEYS)
        w.writeheader()
        for row in data_rows:
            w.writerow({k: row.get(k) for k in BRANDING_FIELD_KEYS})

    from openpyxl import Workbook

    def _write_branding_workbook(path: Path, headers: list[str]) -> None:
        wb_out = Workbook()
        wso = wb_out.active
        # Первый лист Sheet1 — так часто ожидают внешние импортеры; не «subscription_branding».
        wso.title = "Sheet1"
        wso.append(headers)
        for row in data_rows:
            wso.append(
                [
                    row["ID"],
                    row["subscriptionTitle"],
                    row["supportUrl"],
                    row["profileUrl"],
                    row["announcement"],
                    row["active"] == "True",
                ]
            )
        wb_out.save(path)

    _write_branding_workbook(OUT_BRANDING_XLSX, BRANDING_FIELD_KEYS)
    _write_branding_workbook(OUT_BRANDING_XLSX_RU, BRANDING_RU_TITLES)

    print(f"OK -> {OUT_BRANDING_CSV} ({len(data_rows)} rows)")
    print(f"OK -> {OUT_BRANDING_XLSX} ({len(data_rows)} rows, keys + Sheet1)")
    print(f"OK -> {OUT_BRANDING_XLSX_RU} ({len(data_rows)} rows, RU titles + Sheet1)")
    print(
        ">>> NocoBase: import into collection subscription_branding only.",
        "Try",
        OUT_BRANDING_XLSX_RU.name,
        "if cells stay empty (map columns to fields in import wizard).",
        "If you see code / grantDays - wrong collection (Products).",
    )
    return len(data_rows)


def write_products_xlsx(path: Path, import_rows: list[dict[str, object]]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(NOCOBASE_IMPORT_HEADERS)
    for row in import_rows:
        ws.append([row.get(h) for h in NOCOBASE_IMPORT_HEADERS])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print("Установите: pip install openpyxl", file=sys.stderr)
        return 1

    if not XLSX.is_file():
        print(f"Нет файла: {XLSX}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    # Товары: лист «Data», если есть (как в типичном Товары.xlsx), иначе активный лист.
    ws = _find_sheet(wb, "data") or wb.active
    if str(ws.title).strip().lower() == BRANDING_SHEET:
        print(
            "Активный лист — subscription_branding. Откройте в Excel лист с товарами (обычно «Data»), "
            "сохраните файл и запустите скрипт снова.",
            file=sys.stderr,
        )
        return 1
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Пустой лист", file=sys.stderr)
        return 1

    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    def col_idx(*candidates: str) -> int | None:
        lowered = [h.lower() for h in header]
        for cand in candidates:
            c = cand.lower()
            if c in lowered:
                return lowered.index(c)
        return None

    i_code = col_idx("code")
    i_title = col_idx("title")
    i_pt = col_idx("producttype")
    i_so = col_idx("sortorder")
    i_act = col_idx("active")
    i_gd = col_idx("grantdays")
    if i_gd is None:
        # Экспорт NocoBase: подпись типа поля вместо имени
        for j, h in enumerate(header):
            if "целое" in h.lower() and "числ" in h.lower():
                i_gd = j
                break
        if i_gd is None and len(header) > 4:
            i_gd = 4  # как в типичном экспорте: code, затем grantDays

    if i_code is None or i_gd is None:
        print("Не найдены колонки code / grantDays в первой строке.", file=sys.stderr)
        print("Заголовки:", header, file=sys.stderr)
        return 1

    i_srv = col_idx("serverid")

    out_rows: list[dict[str, object]] = []
    for r in rows[1:]:
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        code = r[i_code] if i_code < len(r) else None
        if code is None or str(code).strip() == "":
            continue
        gd = r[i_gd] if i_gd < len(r) else None
        try:
            grant_days = int(float(gd)) if gd is not None and str(gd).strip() != "" else 0
        except (TypeError, ValueError):
            grant_days = 0

        title = ""
        if i_title is not None and i_title < len(r) and r[i_title] is not None:
            title = str(r[i_title]).strip()
        if not title:
            title = f"{grant_days} дней" if grant_days else str(code)

        pt = "vpn_extend"
        if i_pt is not None and i_pt < len(r) and r[i_pt] is not None:
            pt = str(r[i_pt]).strip() or pt

        sort_order = 0
        if i_so is not None and i_so < len(r) and r[i_so] is not None:
            try:
                sort_order = int(float(r[i_so]))
            except (TypeError, ValueError):
                sort_order = 0

        active = True
        if i_act is not None and i_act < len(r) and r[i_act] is not None:
            v = r[i_act]
            if isinstance(v, bool):
                active = v
            else:
                active = str(v).strip().lower() in ("1", "true", "yes", "да")

        server_id = ""
        if i_srv is not None and i_srv < len(r) and r[i_srv] is not None:
            server_id = str(r[i_srv]).strip()

        out_rows.append(
            {
                "code": str(code).strip(),
                "title": title,
                "grantDays": grant_days,
                "productType": pt,
                "sortOrder": sort_order,
                "active": "true" if active else "false",
                "serverId": server_id,
            }
        )

    xlsx_import_rows, csv_import_rows = internal_to_nocobase_import_rows(out_rows)

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=NOCOBASE_IMPORT_HEADERS)
        w.writeheader()
        w.writerows(csv_import_rows)

    write_products_xlsx(OUT_XLSX, xlsx_import_rows)

    print(f"OK -> {OUT_CSV} ({len(out_rows)} rows)")
    print(f"OK -> {OUT_XLSX} ({len(out_rows)} rows)")

    export_subscription_branding(wb)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
